"""
Excel-freies Extraktions-Backend.

Erzeugt dieselben Roh-Artefakte wie der COM-Pfad in :mod:`excel`
(Sheet-CSVs ``Blatt;Adresse;Formel;Wert``, ``vba/<modul>.txt``,
``names_manager.csv``) -- aber ohne Microsoft Excel / pywin32, allein aus der
Datei:

* **openpyxl** liest Formeln (``data_only=False``) und gecachte Werte
  (``data_only=True``) sowie die Defined Names.
* **oletools.olevba** extrahiert den VBA-Quellcode aus ``xl/vbaProject.bin``.

Damit laeuft die Extraktion plattformneutral (Windows/macOS/Linux). Die
nachgelagerte Verarbeitung (Komprimierung, Scalars, Manifest) bleibt
unveraendert, weil das Datei- und Spaltenschema identisch ist.

**Semantik-Hinweis:** ``Wert`` enthaelt den von Excel zuletzt *gespeicherten*
(gecachten) Wert, nicht eine Live-Neuberechnung wie beim COM-Pfad. Fuer
statische, in Excel berechnete Arbeitsmappen ist das aequivalent und sogar
reproduzierbarer (keine Excel-Versions-Varianz). Wurde eine Mappe nie
berechnet/gespeichert, koennen gecachte Werte fehlen.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from rechner_pipeline.extract.excel import (
    _manifest_warning,
    excel_value_to_text,
    is_empty_text,
    safe_filename,
)


def _cell_text(value: Any) -> str:
    """Wandle einen openpyxl-Zellwert in Text fuer das CSV-Schema.

    * Array-/DataTable-Formeln (openpyxl liefert Objekte) -> Formelstring.
    * Ganzzahlige Floats (``5.0``) -> ``"5"`` (entspricht dem COM-Output und
      vermeidet kosmetische Diffs).
    """
    try:
        from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula  # type: ignore

        if isinstance(value, (ArrayFormula, DataTableFormula)):
            value = getattr(value, "text", value)
    except Exception:
        # Fallback ohne harte Abhaengigkeit auf interne openpyxl-Klassen.
        text_attr = getattr(value, "text", None)
        if text_attr is not None and not isinstance(value, (str, int, float, bool)):
            value = text_attr

    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return excel_value_to_text(value)


def _import_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required for the Excel-free export backend. "
            "Install the export dependencies first, e.g. `pip install -e '.[export]'`."
        ) from exc
    return openpyxl


def _import_vba_parser() -> Any:
    try:
        from oletools.olevba import VBA_Parser  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "oletools is required for Excel-free VBA extraction. "
            "Install the export dependencies first, e.g. `pip install -e '.[export]'`."
        ) from exc
    return VBA_Parser


def export_one_sheet(ws_formula, ws_value, sheet_name: str, out_dir: Path) -> Optional[Path]:
    """Schreibe eine Sheet-CSV im Schema ``Blatt;Adresse;Formel;Wert``.

    ``ws_formula`` liefert Formeln/Literale, ``ws_value`` die gecachten Werte
    derselben Zellen. Leere Zellen (weder Formel noch Wert) werden uebersprungen.
    """
    out_path = out_dir / f"{safe_filename(sheet_name)}.csv"
    wrote_any = False

    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Blatt", "Adresse", "Formel", "Wert"])

        for row in ws_formula.iter_rows():
            for cell in row:
                raw_formula = cell.value
                value_cell = ws_value[cell.coordinate]
                raw_value = value_cell.value

                # Formel: Array-Formeln aufloesen, ganzzahlige Literale ohne ".0"
                # (entspricht COMs .Formula). Wert: COMs Float-Repr beibehalten.
                formula_txt = _cell_text(raw_formula)
                value_txt = excel_value_to_text(raw_value)
                if is_empty_text(formula_txt) and is_empty_text(value_txt):
                    continue

                writer.writerow([sheet_name, cell.coordinate, formula_txt, value_txt])
                wrote_any = True

    if not wrote_any:
        try:
            out_path.unlink()
            print(f"[OK] Sheet had no Formel/Wert anywhere -> CSV removed: {sheet_name}")
        except Exception:
            print(f"[OK] Sheet had no Formel/Wert anywhere -> CSV kept: {out_path}")
            return out_path
        return None

    print(f"[OK] Sheet exported (openpyxl): {sheet_name} -> {out_path}")
    return out_path


def export_all_sheets(wb_formula, wb_value, out_dir: Path) -> List[Path]:
    exported: List[Path] = []
    for ws_formula in wb_formula.worksheets:
        ws_value = wb_value[ws_formula.title]
        p = export_one_sheet(ws_formula, ws_value, str(ws_formula.title), out_dir)
        if p is not None and p.exists():
            exported.append(p)
    return exported


def _strip_vba_attribute_lines(code: str) -> str:
    """Entferne ``Attribute VB_*``-Headerzeilen (analog COM ``CodeModule.Lines``)."""
    lines = [ln for ln in code.splitlines() if not ln.startswith("Attribute VB_")]
    return "\n".join(lines)


def export_vba_modules_to_txt(
    excel_path: Path,
    out_dir: Path,
    warnings: List[Dict[str, Any]] | None = None,
) -> List[Path]:
    """Extrahiere VBA-Module via olevba nach ``vba/<modul>.txt``.

    Leere Module (nur Attribut-Header, z. B. Sheet-/Workbook-Klassen ohne Code)
    werden uebersprungen -- wie im COM-Pfad.
    """
    VBA_Parser = _import_vba_parser()

    vba_dir = out_dir / "vba"
    vba_dir.mkdir(parents=True, exist_ok=True)

    exported: List[Path] = []
    parser = None
    try:
        parser = VBA_Parser(str(excel_path))
        if not parser.detect_vba_macros():
            print("[OK] No VBA macros detected.")
        else:
            for _fname, _stream, vba_filename, code in parser.extract_macros():
                module_name = Path(str(vba_filename)).stem
                body = _strip_vba_attribute_lines(str(code))
                if body.strip() == "":
                    continue
                out_path = vba_dir / f"{safe_filename(module_name)}.txt"
                out_path.write_text(body + "\n", encoding="utf-8", newline="\n")
                exported.append(out_path)
                print(f"[OK] VBA exported (olevba): {module_name} -> {out_path}")
    except Exception as exc:
        if warnings is not None:
            warnings.append(
                _manifest_warning(
                    code="export.vba_extraction_failed",
                    stage="export",
                    message="VBA extraction via olevba failed.",
                    strict_error=True,
                    details={"exception": str(exc)},
                )
            )
        print(f"[WARN] VBA extraction failed: {exc}")
    finally:
        if parser is not None:
            try:
                parser.close()
            except Exception:
                pass

    if not exported:
        try:
            vba_dir.rmdir()
        except Exception:
            pass
    return exported


def _scope_for_defined_name(defined_name) -> str:
    if getattr(defined_name, "localSheetId", None) is None:
        return "Workbook"
    return f"Worksheet:{defined_name.localSheetId}"


def _evaluate_single_cell_name(defined_name, wb_value) -> str:
    """Loese den gecachten Wert auf, wenn der Name auf genau eine Zelle zeigt."""
    try:
        destinations = list(defined_name.destinations)
    except Exception:
        return ""
    if len(destinations) != 1:
        return ""
    sheet_title, coord = destinations[0]
    try:
        ws = wb_value[sheet_title]
        return _cell_text(ws[coord.replace("$", "")].value)
    except Exception:
        return ""


def export_name_manager_to_csv(wb_formula, wb_value, out_dir: Path) -> Optional[Path]:
    """Schreibe ``names_manager.csv`` aus den Defined Names (openpyxl)."""
    out_path = out_dir / "names_manager.csv"

    try:
        defined_names = list(wb_formula.defined_names.values())
    except AttributeError:
        # Sehr alte openpyxl-API (definedName-Liste) -- defensiv.
        defined_names = list(getattr(wb_formula.defined_names, "definedName", []))

    rows: List[List[str]] = []
    for dn in defined_names:
        rows.append(
            [
                str(getattr(dn, "name", "")),
                _scope_for_defined_name(dn),
                str(not bool(getattr(dn, "hidden", False))),
                excel_value_to_text(getattr(dn, "value", "")),
                "",  # RefersToLocal: openpyxl liefert keine lokalisierte Form
                "",  # RefersToRangeAddress: keine Live-Range-Aufloesung
                _evaluate_single_cell_name(dn, wb_value),
                "",  # Comment: in DefinedName nicht durchgaengig verfuegbar
            ]
        )

    if not rows:
        print("[OK] No defined names -> no names_manager.csv generated")
        return None

    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(
            [
                "Name",
                "Scope",
                "Visible",
                "RefersTo",
                "RefersToLocal",
                "RefersToRangeAddress",
                "ValueEvaluated",
                "Comment",
            ]
        )
        for row in rows:
            writer.writerow(row)

    print(f"[OK] Name Manager exported (openpyxl) -> {out_path}")
    return out_path


def export_raw(
    excel_path: Path,
    out_dir: Path,
    warnings: List[Dict[str, Any]],
) -> Tuple[List[Path], List[Path], Optional[Path]]:
    """Erzeuge die Roh-Artefakte (Sheet-CSVs, VBA-TXT, Name-Manager) Excel-frei.

    Rueckgabe analog zum COM-Pfad: ``(sheet_csvs, vba_txts, names_manager_csv)``.
    """
    openpyxl = _import_openpyxl()

    wb_formula = openpyxl.load_workbook(excel_path, data_only=False, read_only=False)
    wb_value = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    try:
        sheet_csvs = export_all_sheets(wb_formula, wb_value, out_dir)
        nm_csv = export_name_manager_to_csv(wb_formula, wb_value, out_dir)
    finally:
        wb_formula.close()
        wb_value.close()

    vba_txts = export_vba_modules_to_txt(excel_path, out_dir, warnings=warnings)
    return sheet_csvs, vba_txts, nm_csv
